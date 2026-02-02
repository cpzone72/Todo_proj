import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_bmt_checklist():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NVR BMT Checklist"

    # Define headers
    headers = ["Category", "ID", "Test Item", "Detail Item", "Procedure & Method", "Criteria", "Result (Pass/Fail)", "Remarks"]
    
    # Data structure for the checklist
    checklist_data = [
        # A. General Performance
        ("A. 기본 성능", "A-1", "부팅 속도", "콜드 부팅", "전원 인가 후 라이브/녹화 시작 시간 측정", "3분 이내 녹화 시작", "", ""),
        ("A. 기본 성능", "A-2", "출력 해상도", "4K 듀얼 출력", "HDMI 1/2 독립 4K 모니터 연결 확인", "두 화면 모두 4K@60/30Hz 정상 출력", "", ""),
        ("A. 기본 성능", "A-3", "UI 반응성", "메뉴 조작", "64채널 라이브 중 메뉴 진입/전환 딜레이", "지연(Lag) 없이 즉시 반응", "", ""),

        # B. Processing & Load
        ("B. 영상 처리", "B-1", "입력 대역폭", "최대 대역폭 처리", "64ch, 총 384Mbps 부하 인가", "프레임 드랍 및 영상 깨짐 없음", "", ""),
        ("B. 영상 처리", "B-2", "압축 효율", "Ultra 265 / H.265", "동일 장면 H.264 vs Ultra 265 비트레이트 비교", "약 50% 이상 절감 확인", "", ""),
        ("B. 영상 처리", "B-3", "디코딩 능력", "동시 라이브", "4K 8대 또는 2MP 32대 동시 재생", "끊김 없이 부드러운 재생", "", ""),
        ("B. 영상 처리", "B-4", "장기 부하", "72시간 연속 가동", "64ch 녹화 상태로 3일간 연속 가동", "비정상 재부팅/녹화 누락 0건", "", ""),

        # C. Storage & Reliability
        ("C. 저장/안정성", "C-1", "RAID 구성", "RAID 5/6 구축", "HDD 8개 RAID 볼륨 생성 시간/설정", "정상적인 볼륨 생성 완료", "", ""),
        ("C. 저장/안정성", "C-2", "HDD 장애 대응", "RAID 리빌딩", "녹화 중 HDD 강제 탈거 후 교체/복구", "녹화 중단 없이 리빌딩 수행", "", ""),
        ("C. 저장/안정성", "C-3", "Hot Spare", "N+1 예비 전환", "장애 발생 시 예비 디스크/장비 자동 전환", "설정 시간 내 자동 전환 성공", "", ""),
        ("C. 저장/안정성", "C-4", "ANR 기능", "네트워크 단절 보정", "단선 후 재연결 시 SD카드 데이터 백업", "누락 구간 자동 채움 확인", "", ""),

        # D. AI & Search
        ("D. 지능형 기능", "D-1", "VCA 분석", "침입/라인 크로싱", "구역 설정 후 객체 진입 시 알람 테스트", "검출률 90% 이상, 오알람 최소화", "", ""),
        ("D. 지능형 기능", "D-2", "객체 구분", "사람/차량 분류", "Smart Intrusion Prevention 필터링 동작", "사람/차량 외(동물 등) 무시 확인", "", ""),
        ("D. 지능형 기능", "D-3", "스마트 검색", "AcuSearch", "특정 속성(빨간 옷, 차량) 검색 속도 측정", "수 초 이내 결과 리스트 표출", "", ""),
        ("D. 지능형 기능", "D-4", "얼굴 인식", "Face Detection", "등록된 얼굴 매칭 및 알람 발생 여부", "인식 즉시 알람 표출", "", ""),

        # E. Network & Security
        ("E. 네트워크/보안", "E-1", "원격 접속", "모바일 앱 성능", "외부망(LTE/5G)에서 16분할 로딩 속도", "3초 이내 영상 표출", "", ""),
        ("E. 네트워크/보안", "E-2", "이중화", "듀얼 LAN", "망 분리 또는 로드 밸런싱 설정 테스트", "설정대로 정상 동작 확인", "", ""),
        ("E. 네트워크/보안", "E-3", "보안 프로토콜", "HTTPS/802.1x", "보안 설정 적용 후 영상 전송 상태", "보안 적용 하에 영상 정상 전송", "", ""),

        # F. Compatibility
        ("F. 호환성", "F-1", "타사 연동", "ONVIF 지원", "타사 카메라 등록 및 제어(PTZ 등)", "라이브/녹화/제어 정상 동작", "", "")
    ]

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Write data
    for row_num, row_data in enumerate(checklist_data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            cell.border = thin_border
            # Center align for ID and Result columns
            if col_num in [2, 7]:
                cell.alignment = center_align
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Adjust column widths
    ws.column_dimensions['A'].width = 15  # Category
    ws.column_dimensions['B'].width = 8   # ID
    ws.column_dimensions['C'].width = 15  # Test Item
    ws.column_dimensions['D'].width = 20  # Detail Item
    ws.column_dimensions['E'].width = 40  # Procedure
    ws.column_dimensions['F'].width = 30  # Criteria
    ws.column_dimensions['G'].width = 15  # Result
    ws.column_dimensions['H'].width = 20  # Remarks

    file_name = "NVR3864-V6-IQ_BMT_Checklist.xlsx"
    wb.save(file_name)
    print(f"Excel file saved as {file_name}")

if __name__ == "__main__":
    create_bmt_checklist()
